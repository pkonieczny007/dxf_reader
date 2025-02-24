import os
import openpyxl
import pandas as pd
from shutil import copyfile

# Tworzenie pliku Excel
EXCEL_FILE_1 = '1.WYKAZ_DXF.xlsx'
EXCEL_FILE_2 = '2.PLIK_DO_SPRAWDZENIA.xlsx'
DATA_FILE = 'wykaz.xlsx'

def create_excel_file(file_name, sheet_title):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = sheet_title
    wb.save(file_name)
    wb.close()

def prepare_data(file_path):
    if not os.path.exists("data.csv"):
        if os.path.exists(file_path):
            df = pd.read_excel(file_path)
            df = df[["Nazwa", "Abmess_1", "Abmes_2"]]
            df = df.dropna(subset=["Nazwa"])
            df.to_csv("data.csv", index=False)
            print("Plik data.csv utworzony.")
        else:
            print(f"Plik {file_path} nie istnieje.")

# Funkcja do obliczania szerokości i wysokości
def calculate_dimensions(filename):
    try:
        with open(filename, 'r', encoding='utf-8', errors='replace') as file_obj:
            file_content = file_obj.read()
            data_list = file_content.split()

        min_index = data_list.index('$EXTMIN')
        max_index = data_list.index('$EXTMAX')

        minX = float(data_list[min_index + 2])
        maxX = float(data_list[max_index + 2])

        minY = float(data_list[min_index + 4])
        maxY = float(data_list[max_index + 4])

        width = int(maxX - minX)
        height = int(maxY - minY)

        return width, height
    except Exception as e:
        print(f"Błąd podczas przetwarzania pliku {filename}: {e}")
        return None, None

# Tworzenie wykazu DXF
prepare_data(DATA_FILE)
data = pd.read_csv("data.csv")

# Usuwamy ewentualne spacje w kolumnie "Nazwa" (gdyby były na końcu/początku)
data["Nazwa"] = data["Nazwa"].astype(str).str.strip()  # <-- ZMIANA

wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "wykaz_dxf"

# Nagłówki tabeli
sheet["A1"] = "lp"
sheet["B1"] = "ELEMENT_DXF"
sheet["C1"] = "X-DXF"
sheet["D1"] = "Y-DXF"
sheet["E1"] = "UWAGI"
sheet["F1"] = "Abmess_1"
sheet["G1"] = "Abmes_2"
sheet["H1"] = "SKALA"
sheet["K1"] = "ELEMENT_WYKAZU"

# Przetwarzanie plików DXF
n = 0
for filename in os.listdir():
    if filename.endswith('.dxf'):
        try:
            n += 1
            width, height = calculate_dimensions(filename)

            a = f"A{n+1}"
            b = f"B{n+1}"
            c = f"C{n+1}"
            d = f"D{n+1}"
            e = f"E{n+1}"
            f_ = f"F{n+1}"
            g_ = f"G{n+1}"
            h_ = f"H{n+1}"
            k_ = f"K{n+1}"

            sheet[a] = n

            # ===========================
            # ZMIANA 1: nazwa bez .dxf
            # ===========================
            # Stara wersja:
            # sheet[b] = filename[:-4]
            # Nowa wersja:
            filename_no_ext = os.path.splitext(filename)[0]  # <-- ZMIANA
            sheet[b] = filename_no_ext                        # <-- ZMIANA

            if width is not None and height is not None:
                sheet[c] = width
                sheet[d] = height

                # ===========================
                # ZMIANA 2: dopasowanie do CSV
                # ===========================
                # Stara wersja:
                # base_name = "_".join(filename.split("_")[:4])
                # match = data[data["Nazwa"].str.startswith(base_name)]
                #
                # Nowa wersja:
                base_name = filename_no_ext.strip()           # <-- ZMIANA
                match = data[data["Nazwa"] == base_name]      # <-- ZMIANA

                if not match.empty:
                    abmess_1 = match.iloc[0]["Abmess_1"]
                    abmess_2 = match.iloc[0]["Abmes_2"]
                    nazwa = match.iloc[0]["Nazwa"]

                    sheet[f_] = abmess_1
                    sheet[g_] = abmess_2
                    sheet[k_] = nazwa

                    if width in [abmess_1, abmess_2] and height in [abmess_1, abmess_2]:
                        sheet[h_] = "1:1"
                    else:
                        scale = max(abmess_1, abmess_2) / max(width, height)
                        sheet[h_] = round(scale, 2)
                else:
                    sheet[e] = "Brak w wykazie"
            else:
                sheet[e] = "Błąd"

        except Exception as e:
            print(f"Błąd podczas przetwarzania pliku {filename}: {e}")

# Zapis wyników
wb.save(EXCEL_FILE_1)
wb.close()

# Tworzenie pliku do porównania
copyfile(EXCEL_FILE_1, EXCEL_FILE_2)
wb = openpyxl.load_workbook(EXCEL_FILE_2)
sheet = wb.active
sheet.title = 'porownanie'

# Nagłówki tabeli porównania
sheet['F1'] = 'Abmess_1'
sheet['G1'] = 'Abmes_2'
sheet['H1'] = 'ok/nok - kontrola wymiarow'
sheet['I1'] = 'x =x,y'
sheet['J1'] = 'y =x,y'
sheet['K1'] = 'ELEMENT_WYKAZU'
sheet['L1'] = 'SPRAWDZENIE_NAZWY'
sheet['L2'] = '=B2=K2'

# Wpisywanie formuł
for i in range(2, n + 2):
    sheet[f"L{i}"] = f"=B{i}=K{i}"
    sheet[f"I{i}"] = f"LUB(C{i}=F{i};C{i}=G{i};C{i}=F{i}+1;C{i}=G{i}+1;C{i}=F{i}-1;C{i}=G{i}-1)"
    sheet[f"J{i}"] = f"LUB(D{i}=F{i};D{i}=G{i};D{i}=F{i}+1;D{i}=G{i}+1;D{i}=F{i}-1;D{i}=G{i}-1)"

# Zapis pliku porównania
wb.save(EXCEL_FILE_2)
wb.close()

print(f"Przetworzono {n} plików DXF.")
