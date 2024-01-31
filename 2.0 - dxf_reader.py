import os 
import openpyxl
from shutil import copyfile

#tworzenie pliku excel
wb = openpyxl.Workbook()
wb.save('1.WYKAZ_DXF.xlsx')
wb.close()

#funkcje 
def sze(filename):
    with open(filename, 'r', encoding='utf-8', errors='replace') as file_obj:
        file=file_obj.read()
        lista = file.split()

    min_index= lista.index('$EXTMIN')
    max_index=lista.index('$EXTMAX')

    minx_index=min_index + 2
    maxx_index=max_index +2

    minX=float(lista[minx_index])
    maxX=float(lista[maxx_index])

    miny_index=min_index + 4
    maxy_index=max_index +4          

    minY=float(lista[miny_index])
    maxY=float(lista[maxy_index])

    szerokosc=int(maxX)-int(minX)

    return szerokosc




def wys(filename):
    with open(filename, 'r', encoding='utf-8', errors='replace') as file_obj:
        file=file_obj.read()
        lista = file.split()

    min_index= lista.index('$EXTMIN')
    max_index=lista.index('$EXTMAX')

    minx_index=min_index + 2
    maxx_index=max_index +2

    minX=float(lista[minx_index])
    maxX=float(lista[maxx_index])

    miny_index=min_index + 4
    maxy_index=max_index +4          

    minY=float(lista[miny_index])
    maxY=float(lista[maxy_index])

    wysokosc=int(maxY)-int(minY)

    return wysokosc



#aplikacja tworzenia wykazu dxf

#tworzenie wykazu -twieranie excela
n=0
wb = openpyxl.load_workbook('1.WYKAZ_DXF.xlsx')
sheet = wb.active
sheet.title = 'wykaz_dxf'

#nagłówki tabeli
sheet['A1']='lp'
sheet['B1']='ELEMENT_DXF'
sheet['C1']='X-DXF'
sheet['D1']='Y-DXF'
sheet['E1']='UWAGI'

#tworzenie wykazu
for filename in os.listdir():
    if filename.endswith('.dxf'):
        try:
            n+=1
            wys1=wys(filename)
            szer1=sze(filename)
            print(n, filename, wys1, szer1)

            a="A"+str(n+1)
            b="B"+str(n+1)
            c="C"+str(n+1)
            d="D"+str(n+1)
            
            sheet[a]=str(n)
            sheet[b]=str(filename[:-4])
            sheet[c]=str(wys1)
            sheet[d]=str(szer1)
        except:            
            a="A"+str(n+1)
            b="B"+str(n+1)
            c="C"+str(n+1)
            d="D"+str(n+1)
            e="E"+str(n+1)
            sheet[a]=str(n)
            sheet[b]=str(filename[:-4])
            sheet[c]=str(wys1)
            sheet[d]=str(szer1)
            sheet[e]='blad'

        wb.save('1.WYKAZ_DXF.xlsx')



wb.save('1.WYKAZ_DXF.xlsx')
wb.close()


#tworzenie pliku porównania
src = '1.WYKAZ_DXF.xlsx'
dst = '2.PLIK_DO_SPRAWDZENIA.xlsx'
copyfile(src, dst)


wb = openpyxl.load_workbook('2.PLIK_DO_SPRAWDZENIA.xlsx')
sheet = wb.active
sheet.title = 'porownanie'

#nagłówki tabeli porownania
sheet['F1']='Abmess_1'
sheet['G1']='Abmes_2'
sheet['H1']='ok/nok - kontrola wymiarow'
sheet['I1']='x =x,y'
sheet['J1']='y =x,y'
sheet['K1']='ELEMENT_WYKAZU'
sheet['L1']='SPRAWDZENIE_NAZWY'
sheet['L2']='=B2=K2'


#wpisywanie formuł
for i,j in enumerate(range(n),2):
    x = str(i)
    #formuła - sprawdzanie nazwy
    formulaL = f'=B{i}=K{i}'
    sheet["L"+x]= formulaL

    #formuły - sprawdzanie wymiaru
    formulaI = f'LUB(C{i}=F{i};C{i}=G{i};C{i}=F{i}+1;C{i}=G{i}+1;C{i}=F{i}-1;C{i}=G{i}-1)'
    sheet["I"+x]= formulaI
    
    formulaJ = f'LUB(D{i}=F{i};D{i}=G{i};D{i}=F{i}+1;D{i}=G{i}+1;D{i}=F{i}-1;D{i}=G{i}-1)'
    sheet["J"+x]= formulaJ


#ZAPIS
wb.save('2.PLIK_DO_SPRAWDZENIA.xlsx')
wb.close()
