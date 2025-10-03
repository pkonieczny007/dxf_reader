import os
import openpyxl
import pandas as pd
from shutil import copyfile

# Tworzenie pliku Excel
EXCEL_FILE_1 = '1.WYKAZ_DXF.xlsx'
EXCEL_FILE_2 = '2.PLIK_DO_SPRAWDZENIA.xlsx'
DATA_FILE = 'wykaz.xlsx'

def create_excel_file(file_name, sheet_title):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = sheet_title
    wb.save(file_name)
    wb.close()

def prepare_data(file_path):
    if not os.path.exists("data.csv"):
        if os.path.exists(file_path):
            df = pd.read_excel(file_path)
            df = df[["Nazwa", "Abmess_1", "Abmes_2"]]
            df = df.dropna(subset=["Nazwa"])
            df.to_csv("data.csv", index=False)
            print("Plik data.csv utworzony.")
        else:
            print(f"Plik {file_path} nie istnieje.")

# Funkcja do obliczania szerokości i wysokości
def calculate_dimensions(filename):
    try:
        with open(filename, 'r', encoding='utf-8', errors='replace') as file_obj:
            file_content = file_obj.read()
            data_list = file_content.split()

        min_index = data_list.index('$EXTMIN')
        max_index = data_list.index('$EXTMAX')

        minX = float(data_list[min_index + 2])
        maxX = float(data_list[max_index + 2])

        minY = float(data_list[min_index + 4])
        maxY = float(data_list[max_index + 4])

        width = int(maxX - minX)
        height = int(maxY - minY)

        return width, height
    except Exception as e:
        print(f"Błąd podczas przetwarzania pliku {filename}: {e}")
        return None, None

# Funkcja do dopasowania elementu z wykazu
def find_match_in_data(filename_without_ext, data):
    """
    Najpierw szuka dokładnego dopasowania nazwy,
    jeśli nie znajdzie - szuka po prefiksie (pierwsze 4 segmenty)
    """
    # KROK 1: Szukaj dokładnego dopasowania
    exact_match = data[data["Nazwa"] == filename_without_ext]
    
    if not exact_match.empty:
        print(f"  ✓ Znaleziono dokładne dopasowanie: {filename_without_ext}")
        return exact_match.iloc[0]
    
    # KROK 2: Jeśli nie ma dokładnego, szukaj po prefiksie
    base_name = "_".join(filename_without_ext.split("_")[:4])
    prefix_match = data[data["Nazwa"].str.startswith(base_name)]
    
    if not prefix_match.empty:
        matched_name = prefix_match.iloc[0]["Nazwa"]
        print(f"  ⚠ Dopasowano po prefiksie: {base_name} → {matched_name}")
        return prefix_match.iloc[0]
    
    # KROK 3: Nie znaleziono żadnego dopasowania
    print(f"  ✗ Brak dopasowania dla: {filename_without_ext}")
    return None

# Tworzenie wykazu DXF
prepare_data(DATA_FILE)
data = pd.read_csv("data.csv")
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "wykaz_dxf"

# Nagłówki tabeli
sheet["A1"] = "lp"
sheet["B1"] = "ELEMENT_DXF"
sheet["C1"] = "X-DXF"
sheet["D1"] = "Y-DXF"
sheet["E1"] = "UWAGI"
sheet["F1"] = "Abmess_1"
sheet["G1"] = "Abmes_2"
sheet["H1"] = "SKALA"
sheet["K1"] = "ELEMENT_WYKAZU"

# Przetwarzanie plików DXF
n = 0
print("\n" + "="*60)
print("ROZPOCZYNAM PRZETWARZANIE PLIKÓW DXF")
print("="*60)

for filename in os.listdir():
    if filename.endswith('.dxf'):
        try:
            n += 1
            print(f"\n[{n}] Przetwarzam: {filename}")
            width, height = calculate_dimensions(filename)

            a = f"A{n+1}"
            b = f"B{n+1}"
            c = f"C{n+1}"
            d = f"D{n+1}"
            e = f"E{n+1}"
            f = f"F{n+1}"
            g = f"G{n+1}"
            h = f"H{n+1}"
            k = f"K{n+1}"

            sheet[a] = n
            filename_without_ext = filename[:-4]
            sheet[b] = filename_without_ext

            if width is not None and height is not None:
                sheet[c] = width
                sheet[d] = height

                # Dopasowanie danych z wykazu - POPRAWIONA WERSJA
                match = find_match_in_data(filename_without_ext, data)

                if match is not None:
                    abmess_1 = match["Abmess_1"]
                    abmess_2 = match["Abmes_2"]
                    nazwa = match["Nazwa"]

                    sheet[f] = abmess_1
                    sheet[g] = abmess_2
                    sheet[k] = nazwa

                    if width in [abmess_1, abmess_2] and height in [abmess_1, abmess_2]:
                        sheet[h] = "1:1"
                    else:
                        scale = max(abmess_1, abmess_2) / max(width, height)
                        sheet[h] = round(scale, 2)
                else:
                    sheet[e] = "Brak w wykazie"
            else:
                sheet[e] = "Błąd"

        except Exception as e:
            print(f"  ✗ Błąd podczas przetwarzania pliku {filename}: {e}")

# Zapis wyników
wb.save(EXCEL_FILE_1)
wb.close()

print("\n" + "="*60)
print(f"✓ Przetworzono {n} plików DXF.")
print(f"✓ Zapisano wyniki w: {EXCEL_FILE_1}")
print("="*60)

# Tworzenie pliku do porównania
print("\nTworzę plik do sprawdzenia...")

# Wczytaj dane z pierwszego pliku
wb_source = openpyxl.load_workbook(EXCEL_FILE_1)
sheet_source = wb_source.active

# Utwórz nowy workbook
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = 'porownanie'

# Skopiuj wszystkie dane z pierwszego pliku
for row in sheet_source.iter_rows():
    for cell in row:
        new_cell = sheet[cell.coordinate]
        new_cell.value = cell.value
        if cell.has_style:
            new_cell.font = cell.font.copy()
            new_cell.border = cell.border.copy()
            new_cell.fill = cell.fill.copy()
            new_cell.number_format = cell.number_format
            new_cell.protection = cell.protection.copy()
            new_cell.alignment = cell.alignment.copy()

wb_source.close()

# Nagłówki tabeli porównania
sheet['F1'] = 'Abmess_1'
sheet['G1'] = 'Abmes_2'
sheet['H1'] = 'ok/nok - kontrola wymiarow'
sheet['I1'] = 'x =x,y'
sheet['J1'] = 'y =x,y'
sheet['K1'] = 'ELEMENT_WYKAZU'
sheet['L1'] = 'SPRAWDZENIE_NAZWY'

# Wpisywanie formuł
for i in range(2, n + 2):
    sheet[f"L{i}"].value = f"=B{i}=K{i}"
    sheet[f"I{i}"].value = f"=OR(C{i}=F{i},C{i}=G{i},C{i}=F{i}+1,C{i}=G{i}+1,C{i}=F{i}-1,C{i}=G{i}-1)"
    sheet[f"J{i}"].value = f"=OR(D{i}=F{i},D{i}=G{i},D{i}=F{i}+1,D{i}=G{i}+1,D{i}=F{i}-1,D{i}=G{i}-1)"

# Zapis pliku porównania
wb.save(EXCEL_FILE_2)
wb.close()

print(f"✓ Zapisano plik do sprawdzenia: {EXCEL_FILE_2}")
print("\n" + "="*60)
print("ZAKOŃCZONO POMYŚLNIE")
print("="*60)
